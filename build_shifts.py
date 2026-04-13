import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "2026 Roster"

# 1. STRICT 8-CHARACTER ARGB COLORS (This fixes the "Red Screen" bug!)
DAY_FILL = "FFFFF2CC"     # Light Yellow
NIGHT_FILL = "FFDDEBF7"   # Light Blue
WEEKEND_GRAY = "FFF2F2F2" # Light Gray
HEADER_GRAY = "FFE7E6E6"  # Darker Gray
SAP_RED = "FFFF7C80"
PLAN_YELLOW = "FFFFFF00"

DAY_FONT = "FFD6B656"
NIGHT_FONT = "FF6C8EBF"

# 2. PROFESSIONAL CORNERING RULES (Borders)
thin_border = Side(border_style="thin", color="FFD9D9D9")
grid_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)

# 3. Setup the 4 Anchor Columns
headers = ["Emp ID", "Full Name", "Crew", "Role"]
for col, text in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=text)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color=HEADER_GRAY, end_color=HEADER_GRAY, fill_type="solid")
    cell.border = grid_border # Apply cornering
    ws.column_dimensions[get_column_letter(col)].width = 16

ws.column_dimensions['B'].width = 32 

# 4. Setup the 365-Day Calendar Grid Headers
start_date = date(2026, 1, 1)
months = ["January", "February", "March", "April", "May", "June", 
          "July", "August", "September", "October", "November", "December"]
month_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

col_idx = 5 
for m_idx, m_name in enumerate(months):
    days_in_month = month_days[m_idx]
    
    ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + days_in_month - 1)
    m_cell = ws.cell(row=2, column=col_idx, value=m_name)
    m_cell.font = Font(bold=True)
    m_cell.alignment = Alignment(horizontal="center", vertical="center")
    m_cell.fill = PatternFill(start_color=HEADER_GRAY, end_color=HEADER_GRAY, fill_type="solid")
    m_cell.border = grid_border

    for d in range(days_in_month):
        current_col = col_idx + d
        col_letter = get_column_letter(current_col)
        
        ws.column_dimensions[col_letter].width = 3.5 
        
        d_cell = ws.cell(row=3, column=current_col, value=d+1)
        d_cell.alignment = Alignment(horizontal="center")
        d_cell.border = grid_border
        
        current_date = start_date + timedelta(days=(current_col - 5))
        dow_cell = ws.cell(row=4, column=current_col, value=current_date.strftime("%a")[0]) 
        dow_cell.alignment = Alignment(horizontal="center")
        dow_cell.font = Font(size=8, color="FF595959")
        dow_cell.border = grid_border
        
        # Highlight headers for Friday/Saturday
        if current_date.weekday() in [4, 5]: 
            d_cell.fill = PatternFill(start_color=WEEKEND_GRAY, end_color=WEEKEND_GRAY, fill_type="solid")
            dow_cell.fill = PatternFill(start_color=WEEKEND_GRAY, end_color=WEEKEND_GRAY, fill_type="solid")

    col_idx += days_in_month

# 5. Add Legend
ws.merge_cells('A1:D1')
title = ws['A1']
title.value = "2026 OPERATIONS & LEAVE ROSTER"
title.font = Font(size=14, bold=True)

ws['E1'] = " S = SAP"
ws['E1'].fill = PatternFill(start_color=SAP_RED, end_color=SAP_RED, fill_type="solid")
ws['E1'].font = Font(bold=True)
ws['E1'].border = grid_border
ws.merge_cells('E1:F1')

ws['H1'] = " P = Plan"
ws['H1'].fill = PatternFill(start_color=PLAN_YELLOW, end_color=PLAN_YELLOW, fill_type="solid")
ws['H1'].font = Font(bold=True)
ws['H1'].border = grid_border
ws.merge_cells('H1:I1')

ws['K1'] = " D = Day Shift"
ws['K1'].fill = PatternFill(start_color=DAY_FILL, end_color=DAY_FILL, fill_type="solid")
ws['K1'].border = grid_border
ws.merge_cells('K1:M1')

ws['O1'] = " N = Night Shift"
ws['O1'].fill = PatternFill(start_color=NIGHT_FILL, end_color=NIGHT_FILL, fill_type="solid")
ws['O1'].border = grid_border
ws.merge_cells('O1:R1')

# 6. Inject Crew Data, Names, and Shifts
shift_cycles = {
    'A': ['O', 'O', 'O', 'O', 'D', 'D', 'N', 'N'],
    'B': ['D', 'D', 'N', 'N', 'O', 'O', 'O', 'O'],
    'C': ['N', 'N', 'O', 'O', 'O', 'O', 'D', 'D'],
    'D': ['O', 'O', 'D', 'D', 'N', 'N', 'O', 'O'],
    'General': ['O', 'O', 'O', 'O', 'O', 'O', 'O', 'O'] 
}

crews_data = {
    'A': {
        'Shift In Charge': [("502667", "Zaid Almarri")],
        'Supervisor': [("2025", "Syed Shahnawaz Ahmed")],
        'CCR Operator': [("2364", "Sami Hamdan"), ("2954", "Fawaz Al Qahtani"), ("2726", "Ahmed Fathy"), ("2202", "Abdulwahab Al-Shehab"), ("2711", "Shaheer Yousaf")],
        'Local Operator': [("500438", "Purushothaman Devaraj"), ("3865", "Mohamed Al Dossary"), ("501369", "Abdullah Altulayhi"), ("3617", "Bakr Abdulmajeed"), ("4101", "Saad Salem Al Hajiri"), ("4095", "Abdulhadi Mohammed Saleh")]
    },
    'B': {
        'Shift In Charge': [("2363", "Abdullah Al Amri")],
        'Supervisor': [("500446", "Kanaka Naga Srinivasu Kolli")],
        'CCR Operator': [("3672", "Prathapan"), ("", "Ahmed Fawaz"), ("2369", "Rashed Alhajri"), ("2035", "Saravanakumar Madhaiyan"), ("3251", "Abdullah Mohiddin Alshahrani")],
        'Local Operator': [("3262", "Mohammed Mulhim"), ("500439", "Mohammed Afnan"), ("3511", "Abdelaziz Al Harbi"), ("4091", "Amjad Alruwaytie"), ("4098", "Abderahman Al Anezie")]
    },
    'C': {
        'Shift In Charge': [("502680", "Alaa Alrefae")],
        'Supervisor': [("2038", "Moustafa Elansari")],
        'CCR Operator': [("2665", "Abdullah Al Hajri"), ("501370", "Yasser Althuiqeb"), ("501388", "Waleed Fayad"), ("2283", "Norbie Cruz"), ("2028", "Juma Khan")],
        'Local Operator': [("501885", "Khalid Saleh Khalousi"), ("3229", "Saad Al Enize"), ("500440", "Izhar Ali"), ("4096", "Abderahman Al Baqmi"), ("4108", "Ali Al Qahtani")]
    },
    'D': {
        'Shift In Charge': [("2024", "Mustafa Salem")],
        'Supervisor': [],
        'CCR Operator': [("501234", "Lakshmy Prasad"), ("2941", "Saleh Mohammed Al Amri"), ("2512", "Hassan Arshad"), ("2294", "Veera Prasad")],
        'Local Operator': [("2492", "Rajesh Muniasamy"), ("2237", "Mark Ramirez"), ("2820", "Mohammed Al Ghamdi"), ("2912", "Bader AlSubait"), ("4097", "Faris Al Dawasri"), ("4081", "Mohamed Al Hakami"), ("4094", "Saad Al Shahrani"), ("3973", "Ahmed Alsaqoor")]
    },
    'General': {
        'Supervisor': [("1119", "Abdul Hameed")],
        'CCR Operator': [("3142", "Mohammad Algarni")]
    }
}

# Strict 8-char ARGB name colors
name_colors = {
    'Syed Shahnawaz Ahmed': 'FFFF0000', 'Abdullah Al Amri': 'FFFF0000', 'Kanaka Naga Srinivasu Kolli': 'FFFF0000', 'Moustafa Elansari': 'FFFF0000', 'Mustafa Salem': 'FFFF0000', 'Zaid Almarri': 'FFFF0000', 'Alaa Alrefae': 'FFFF0000',
    'Saleh Mohammed Al Amri': 'FFD4AF37', 'Adam Mohammed Al-Huzum': 'FFD4AF37', 'Sami Hamdan': 'FFD4AF37',
    'Shaheer Yousaf': 'FF00B050', 'Juma Khan': 'FF00B050', 'Veera Prasad': 'FF00B050', 'Waleed Fayad': 'FF00B050',
    'Ahmed Fathy': 'FF0070C0', 'Fawaz Al Qahtani': 'FF0070C0', 'Rashed Alhajri': 'FF0070C0', 'Prathapan': 'FF0070C0', 'Yasser Althuiqeb': 'FF0070C0', 'Abdullah Al Hajri': 'FF0070C0', 'Norbie Cruz': 'FF0070C0', 'Lakshmy Prasad': 'FF0070C0', 'Saravanakumar Madhaiyan': 'FF0070C0', 'Hassan Arshad': 'FF0070C0', 'Ahmed Fawaz': 'FF0070C0',
    'Purushothaman Devaraj': 'FF7030A0', 'Abdulwahab Al-Shehab': 'FF7030A0', 'Izhar Ali': 'FF7030A0', 'Mark Ramirez': 'FF7030A0', 'Ahmed Alsaqoor': 'FF7030A0', 'Rajesh Muniasamy': 'FF7030A0',
    'Mohamed Al Dossary': 'FFED7D31', 'Abdullah Altulayhi': 'FFED7D31', 'Bakr Abdulmajeed': 'FFED7D31', 'Mohammed Mulhim': 'FFED7D31', 'Abdullah Mohiddin Alshahrani': 'FFED7D31', 'Mohammed Afnan': 'FFED7D31', 'Abdelaziz Al Harbi': 'FFED7D31', 'Amjad Alruwaytie': 'FFED7D31', 'Khalid Saleh Khalousi': 'FFED7D31', 'Saad Al Enize': 'FFED7D31', 'Mohammed Al Ghamdi': 'FFED7D31', 'Bader AlSubait': 'FFED7D31',
    'Saad Salem Al Hajiri': 'FF808080', 'Abdulhadi Mohammed Saleh': 'FF808080', 'Abderahman Al Anezie': 'FF808080', 'Abderahman Al Baqmi': 'FF808080', 'Ali Al Qahtani': 'FF808080', 'Faris Al Dawasri': 'FF808080', 'Mohamed Al Hakami': 'FF808080', 'Saad Al Shahrani': 'FF808080'
}

row_num = 5

for crew_name, roles in crews_data.items():
    cycle = shift_cycles[crew_name]
    
    for role_name, employees in roles.items():
        for emp_id, emp_name in employees:
            
            # Write Anchors with Borders
            c1 = ws.cell(row=row_num, column=1, value=emp_id)
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.border = grid_border
            
            c2 = ws.cell(row=row_num, column=2, value=emp_name)
            c2.alignment = Alignment(vertical="center")
            c2.border = grid_border
            if emp_name in name_colors:
                c2.font = Font(color=name_colors[emp_name], bold=True)
            else:
                c2.font = Font(bold=True)
                
            c3 = ws.cell(row=row_num, column=3, value=f"Crew {crew_name}" if crew_name != 'General' else "General")
            c3.alignment = Alignment(horizontal="center", vertical="center")
            c3.border = grid_border
            
            c4 = ws.cell(row=row_num, column=4, value=role_name)
            c4.alignment = Alignment(horizontal="center", vertical="center")
            c4.border = grid_border
            
            # Map the 365 days
            for d in range(365):
                col_num = 5 + d
                shift = cycle[d % 8] 
                
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = grid_border # PROFESSIONAL CORNERING applied to every day block!
                
                current_date = start_date + timedelta(days=d)
                is_weekend = current_date.weekday() in [4, 5]

                if shift == 'D':
                    cell.value = "D"
                    cell.fill = PatternFill(start_color=DAY_FILL, end_color=DAY_FILL, fill_type="solid")
                    cell.font = Font(color=DAY_FONT, bold=True)
                elif shift == 'N':
                    cell.value = "N"
                    cell.fill = PatternFill(start_color=NIGHT_FILL, end_color=NIGHT_FILL, fill_type="solid")
                    cell.font = Font(color=NIGHT_FONT, bold=True)
                else:
                    # Brings back the missing weekend colors!
                    if is_weekend: 
                        cell.fill = PatternFill(start_color=WEEKEND_GRAY, end_color=WEEKEND_GRAY, fill_type="solid")

            row_num += 1

# 7. Freeze Panes
ws.freeze_panes = "E5"

wb.save("Perfect_Master_Template_2026.xlsx")
print("✅ Created Perfect_Master_Template_2026.xlsx successfully with Fixed Colors and Professional Cornering!")